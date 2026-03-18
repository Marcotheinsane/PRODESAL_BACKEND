#!/usr/bin/env python
"""
Diagnóstico de la estructura del Excel
Muestra: hojas, columnas, primeras filas para identificar la estructura
"""
import openpyxl
import os

excel_file = 'NOMINA 2025 CON SECTOR.xlsx'

if not os.path.exists(excel_file):
    print(f"Error: {excel_file} no encontrado")
    exit(1)

print(f"\n{'='*80}")
print(f"DIAGNÓSTICO DE ESTRUCTURA: {excel_file}")
print(f"{'='*80}\n")

wb = openpyxl.load_workbook(excel_file)

print(f"📄 Hojas encontradas: {len(wb.sheetnames)}")
for i, sheet in enumerate(wb.sheetnames, 1):
    print(f"   {i}. {sheet}")

print(f"\n{'='*80}")
print(f"ANÁLISIS DE LA PRIMERA HOJA")
print(f"{'='*80}\n")

ws = wb.active
print(f"Hoja activa: {ws.title}")
print(f"Dimensiones: {ws.dimensions}")

print(f"\n📋 ESTRUCTURA DE COLUMNAS (Fila 1):")
print("-" * 80)
headers = []
for col_idx, cell in enumerate(ws[1], 1):
    value = cell.value
    headers.append(value)
    print(f"   Columna {col_idx:2} ({chr(64+col_idx)}): {value}")

print(f"\n📊 PRIMERAS 10 FILAS DE DATOS:")
print("-" * 80)

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=False), start=2):
    print(f"\nFila {row_idx}:")
    for col_idx, cell in enumerate(row[:10], 1):  # Primeras 10 columnas
        value = cell.value
        if value is not None:
            # Si es una fecha, mostrar su tipo
            if hasattr(value, 'date'):
                print(f"   {chr(64+col_idx)}: {value} (tipo: datetime)")
            else:
                print(f"   {chr(64+col_idx)}: {value} (tipo: {type(value).__name__})")

print(f"\n{'='*80}")
print(f"CONTEO DE DATOS POR COLUMNA")
print(f"{'='*80}\n")

# Contar filas con datos en cada columna
for col_idx in range(1, len(headers) + 1):
    col_data = ws[openpyxl.utils.get_column_letter(col_idx)]
    non_empty = sum(1 for cell in col_data if cell.value is not None)
    print(f"   Columna {openpyxl.utils.get_column_letter(col_idx)}: {non_empty} celdas con datos")

print(f"\n{'='*80}")
print("✓ Diagnóstico completado")
print(f"{'='*80}\n")
