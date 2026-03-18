from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
from openpyxl import load_workbook
from io import BytesIO
from django.db import models
from django.db.models import Q

#models
from .models import Clientes, Asunto, RegistroAsunto, Asistencia
from .serializers import (
    ClientesListSerializer, ClientesDetailSerializer, 
    AsuntoSerializer, RegistroAsuntoSerializer, AsistenciaSerializer
)

#filtros
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters

# Create your views here.

class UploadExcelView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)
    
    def _formatear_rut(self, rut):
        if not rut:
            return None
        
        # Convertir a string
        rut_str = str(rut).strip()
        
        # Si ya tiene el formato XX.XXX.XXX-X, devolverlo tal cual
        if '.' in rut_str and '-' in rut_str:
            return rut_str
        
        # Si no, limpiar y formatear
        rut_limpio = rut_str.replace('.', '').replace('-', '').strip()
        
        # Debe tener al menos 8 caracteres (7 dígitos + 1 verificador)
        if len(rut_limpio) < 8:
            return None
        
        # Separar dígito verificador (último carácter)
        cuerpo = rut_limpio[:-1]
        digito = rut_limpio[-1]
        
        # Formatear: XX.XXX.XXX-X
        # Padronizar tamaño: si tiene menos de 7 dígitos, agregar ceros al inicio
        while len(cuerpo) < 7:
            cuerpo = '0' + cuerpo
        
        # Formato final: primeros 2 dígitos, punto, siguientes 3, punto, siguientes 2, guión, dígito verificador
        rut_formateado = f"{cuerpo[0:2]}.{cuerpo[2:5]}.{cuerpo[5:7]}-{digito}"
        return rut_formateado
    
    def post(self, request):
        """Subir archivo Excel con clientes"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        # Validar que sea Excel
        if not file.name.endswith(('.xlsx', '.xls', '.csv')):
            return Response(
                {'error': 'El archivo debe ser Excel (.xlsx, .xls) o CSV'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Leer Excel
            archivo = BytesIO(file.read())
            workbook = load_workbook(archivo)
            worksheet = workbook.active
            
            creados = 0
            errores = []
            
            # Procesar filas (EMPEZAR DESDE FILA 1 porque no hay encabezado)
            for idx, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True), start=1):
                try:
                    # Las columnas están desplazadas porque la columna A tiene números de índice
                    # Estructura: A(índice), B(RUT), C(Nombre1), D(Nombre2), E(Sector), F(?)
                    rut = row[1]  # Columna B
                    nombres = row[2]  # Columna C
                    apellidos = row[3]  # Columna D
                    sector = row[4] if len(row) > 4 else None  # Columna E
                    
                    # Debug: Log primeras 5 filas
                    if idx <= 6:
                        print(f"FILA {idx}: RUT='{rut}', NOMBRES='{nombres}', APELLIDOS='{apellidos}', SECTOR='{sector}'")
                    
                    # Formatear RUT
                    rut = self._formatear_rut(rut)
                    
                    # Validar datos requeridos
                    if not rut or not nombres or not apellidos:
                        errores.append({
                            'fila': idx,
                            'error': 'Faltan RUT, NOMBRES o APELLIDOS (RUT debe tener formato válido)'
                        })
                        continue
                    
                    # Crear o actualizar cliente
                    cliente, created = Clientes.objects.get_or_create(
                        rut=rut,
                        defaults={
                            'nombres': str(nombres),
                            'apellidos': str(apellidos),
                            'sector': str(sector) if sector else None,
                            'es_beneficiario': False
                        }
                    )
                    
                    if created:
                        creados += 1
                    else:
                        # Actualizar si ya existe
                        cliente.nombres = str(nombres)
                        cliente.apellidos = str(apellidos)
                        cliente.sector = str(sector) if sector else None
                        cliente.save()
                        creados += 1
                        
                except Exception as e:
                    errores.append({
                        'fila': idx,
                        'error': str(e)
                    })
            
            return Response({
                'mensaje': f'Importación completada',
                'creados': creados,
                'errores': errores,
                'total_errores': len(errores)
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error al procesar Excel: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

class ClientesViewSet(viewsets.ModelViewSet):
    """
    API ViewSet para gestionar clientes del programa INDAP con fichas sociales completas.
    
    Endpoints disponibles:
    - GET /api/v1/clientes/ - Listar todos los clientes
    - POST /api/v1/clientes/ - Crear nuevo cliente
    - GET /api/v1/clientes/{id}/ - Obtener detalle completo de cliente
    - PUT/PATCH /api/v1/clientes/{id}/ - Actualizar cliente
    - DELETE /api/v1/clientes/{id}/ - Eliminar cliente
    - GET /api/v1/clientes/por-rut/{rut}/ - Buscar por RUT (con ficha social completa)
    - POST /api/v1/clientes/upload-excel/ - Subir clientes desde Excel
    """
    queryset = Clientes.objects.prefetch_related('hogar__carga_familiar', 'informacion_salud', 'redes_conocidas', 'registro_social').all()
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['sector', 'es_beneficiario']
    search_fields = ['nombres', 'apellidos', 'rut', 'telefono']
    ordering_fields = ['fecha_postulacion', 'nombres', 'rut']
    ordering = ['-fecha_postulacion']
    
    def get_serializer_class(self):
        """Usar serializer diferente según la acción"""
        if self.action == 'retrieve' or self.basename == 'cliente-por-rut':
            return ClientesDetailSerializer
        return ClientesListSerializer
    
    @action(detail=False, methods=['get'], url_path='por-rut/(?P<rut>[^/.]+)')
    def por_rut(self, request, rut=None):
        """
        Obtener cliente y su ficha social completa por RUT
        GET /api/v1/clientes/por-rut/{rut}/
        
        Incluye:
        - Información personal completa
        - Hogar y datos de vivienda
        - Carga familiar (hijos, padres, etc)
        - Información de salud
        - Redes de apoyo
        - Registro Social de Hogares
        - Necesidades, observaciones y recomendaciones
        """
        try:
            cliente = Clientes.objects.prefetch_related(
                'hogar__carga_familiar',
                'informacion_salud',
                'redes_conocidas',
                'registro_social'
            ).get(rut=rut)
            serializer = ClientesDetailSerializer(cliente)
            return Response(serializer.data)
        except Clientes.DoesNotExist:
            return Response(
                {'error': f'Cliente con RUT {rut} no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=False, methods=['get'])
    def buscar(self, request):
        """
        Búsqueda avanzada de clientes
        GET /api/v1/clientes/buscar/?q=término&sector=sector&beneficiario=true
        """
        q = request.query_params.get('q', '')
        sector = request.query_params.get('sector', '')
        beneficiario = request.query_params.get('beneficiario', '')
        
        queryset = self.queryset
        
        if q:
            queryset = queryset.filter(
                Q(rut__icontains=q) |
                Q(nombres__icontains=q) |
                Q(apellidos__icontains=q) |
                Q(telefono__icontains=q)
            )
        
        if sector:
            queryset = queryset.filter(sector__icontains=sector)
        
        if beneficiario.lower() in ('true', 'si', '1', 'yes'):
            queryset = queryset.filter(es_beneficiario=True)
        elif beneficiario.lower() in ('false', 'no', '0'):
            queryset = queryset.filter(es_beneficiario=False)
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(
        detail=False, 
        methods=['post'], 
        permission_classes=[IsAuthenticated],
        parser_classes=(MultiPartParser, FormParser)
    )
    def upload_excel(self, request):
        """
        Subir archivo Excel con clientes
        POST /api/v1/clientes/upload-excel/
        
        El Excel debe tener columnas:
        - A: RUT (formato: XX.XXX.XXX-X)
        - B: NOMBRES
        - C: APELLIDOS
        - D: SECTOR
        - E: EXTENS (opcional)
        - F: ES_BENEFICIARIO (opcional, True/False)
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        # Validar que sea Excel
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'El archivo debe ser Excel (.xlsx o .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Leer Excel
            archivo = BytesIO(file.read())
            workbook = load_workbook(archivo)
            worksheet = workbook.active
            
            creados = 0
            errores = []
            
            # Procesar filas (empezar desde fila 2 para saltar encabezado)
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    rut = row[0]
                    nombres = row[1]
                    apellidos = row[2]
                    sector = row[3] if len(row) > 3 else None
                    extens = row[4] if len(row) > 4 else None
                    es_beneficiario = row[5] if len(row) > 5 else False
                    
                    # Validar datos requeridos
                    if not rut or not nombres or not apellidos:
                        errores.append({
                            'fila': idx,
                            'error': 'Faltan RUT, NOMBRES o APELLIDOS'
                        })
                        continue
                    
                    # Convertir es_beneficiario a booleano
                    if isinstance(es_beneficiario, str):
                        es_beneficiario = es_beneficiario.lower() in ('true', 'sí', 'si', '1', 'yes')
                    else:
                        es_beneficiario = bool(es_beneficiario)
                    
                    # Crear o actualizar cliente
                    cliente, created = Clientes.objects.get_or_create(
                        rut=str(rut),
                        defaults={
                            'nombres': str(nombres),
                            'apellidos': str(apellidos),
                            'sector': str(sector) if sector else None,
                            'extens': str(extens) if extens else None,
                            'es_beneficiario': es_beneficiario
                        }
                    )
                    
                    if created:
                        creados += 1
                    else:
                        # Actualizar si ya existe
                        cliente.nombres = str(nombres)
                        cliente.apellidos = str(apellidos)
                        cliente.sector = str(sector) if sector else None
                        cliente.extens = str(extens) if extens else None
                        cliente.es_beneficiario = es_beneficiario
                        cliente.save()
                        creados += 1
                        
                except Exception as e:
                    errores.append({
                        'fila': idx,
                        'error': str(e)
                    })
            
            return Response({
                'mensaje': f'Importación completada',
                'creados': creados,
                'errores': errores,
                'total_errores': len(errores)
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error al procesar Excel: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def beneficiarios(self, request):
        """
        Endpoint especial para obtener solo beneficiarios
        GET /api/v1/clientes/beneficiarios/
        """
        queryset = self.queryset.filter(es_beneficiario=True)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def por_sector(self, request):
        """
        Endpoint para obtener estadísticas por sector
        GET /api/v1/clientes/por_sector/
        """
        sectores = Clientes.objects.values('sector').distinct()
        data = []
        for sector_obj in sectores:
            sector = sector_obj['sector']
            if sector:
                count = Clientes.objects.filter(sector=sector).count()
                beneficiarios = Clientes.objects.filter(sector=sector, es_beneficiario=True).count()
                data.append({
                    'sector': sector,
                    'total': count,
                    'beneficiarios': beneficiarios,
                    'no_beneficiarios': count - beneficiarios
                })
        return Response(data)


class AsuntoViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar Asuntos/Eventos
    """
    queryset = Asunto.objects.all()
    serializer_class = AsuntoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre', 'tipo']
    ordering_fields = ['nombre', 'total_instancias', '-updated_at']
    ordering = ['-updated_at']
    
    @action(detail=True, methods=['get'], permission_classes=[IsAuthenticated])
    def historial(self, request, pk=None):
        """
        Obtener historial cronológico del asunto
        GET /api/v1/asuntos/{id}/historial/
        """
        asunto = self.get_object()
        registros = asunto.registros.all()
        serializer = RegistroAsuntoSerializer(registros, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'], permission_classes=[IsAuthenticated])
    def asistentes(self, request, pk=None):
        """
        Obtener asistentes de todas las instancias del asunto
        GET /api/v1/asuntos/{id}/asistentes/
        """
        asunto = self.get_object()
        registros = asunto.registros.all()
        
        data = []
        for registro in registros:
            asistencias = registro.asistencias.all()
            for asistencia in asistencias:
                data.append({
                    'id': asistencia.id,
                    'cliente_id': asistencia.cliente.id,
                    'cliente_rut': asistencia.cliente.rut,
                    'cliente_nombres': asistencia.cliente.nombres,
                    'cliente_apellidos': asistencia.cliente.apellidos,
                    'registro_asunto_id': registro.id,
                    'fecha_evento': registro.fecha,
                    'presente': asistencia.presente,
                })
        
        return Response(data)


class RegistroAsuntoViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar Registros de Asuntos (instancias específicas)
    """
    queryset = RegistroAsunto.objects.all()
    serializer_class = RegistroAsuntoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['asunto']
    ordering_fields = ['fecha', '-fecha']
    ordering = ['-fecha']


class AsistenciaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar Asistencias
    """
    queryset = Asistencia.objects.all()
    serializer_class = AsistenciaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['cliente', 'registro_asunto', 'presente']
    
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def registrar_asistentes(self, request):
        """
        Registrar múltiples asistentes a la vez
        POST /api/v1/asistencias/registrar_asistentes/
        {
            "registro_asunto_id": 1,
            "asistentes": [
                {"cliente_id": 1, "presente": true},
                {"cliente_id": 2, "presente": false}
            ]
        }
        """
        registro_asunto_id = request.data.get('registro_asunto_id')
        asistentes = request.data.get('asistentes', [])
        
        if not registro_asunto_id:
            return Response(
                {'error': 'Se requiere registro_asunto_id'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            registro_asunto = RegistroAsunto.objects.get(id=registro_asunto_id)
        except RegistroAsunto.DoesNotExist:
            return Response(
                {'error': 'Registro de asunto no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        creados = 0
        actualizados = 0
        errores = []
        
        for asistente in asistentes:
            try:
                cliente_id = asistente.get('cliente_id')
                presente = asistente.get('presente', False)
                
                if not cliente_id:
                    errores.append({'error': 'Falta cliente_id'})
                    continue
                
                asistencia, created = Asistencia.objects.get_or_create(
                    cliente_id=cliente_id,
                    registro_asunto=registro_asunto,
                    defaults={'presente': presente}
                )
                
                if not created:
                    asistencia.presente = presente
                    asistencia.save()
                    actualizados += 1
                else:
                    creados += 1
                    
            except Exception as e:
                errores.append({'error': str(e)})
        
        return Response({
            'creados': creados,
            'actualizados': actualizados,
            'errores': errores
        })