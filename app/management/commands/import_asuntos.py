from django.core.management.base import BaseCommand
from django.db import IntegrityError
from openpyxl import load_workbook
from datetime import datetime
from app.models import Asunto, RegistroAsunto
from collections import defaultdict


class Command(BaseCommand):
    help = 'Importa asuntos desde un archivo Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            'archivo',
            type=str,
            help='Ruta del archivo Excel con los asuntos'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default='Asuntos',
            help='Nombre de la hoja Excel (default: "Asuntos")'
        )

    def handle(self, *args, **options):
        archivo = options['archivo']
        nombre_hoja = options['sheet']

        try:
            # Cargar workbook
            wb = load_workbook(archivo)
            
            # Seleccionar hoja
            if nombre_hoja not in wb.sheetnames:
                self.stdout.write(
                    self.style.ERROR(f'Error: Hoja "{nombre_hoja}" no encontrada.')
                )
                self.stdout.write(f'Hojas disponibles: {", ".join(wb.sheetnames)}')
                return

            ws = wb[nombre_hoja]

            # Procesar datos
            asuntos_map = defaultdict(list)  # {nombre: [registro1, registro2, ...]}
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=False), 1):
                # Saltar filas vacías
                if not any(cell.value for cell in row):
                    continue

                try:
                    # Obtener valores
                    nombre = row[0].value
                    tipo = row[1].value if len(row) > 1 else None
                    fecha_str = row[2].value if len(row) > 2 else None
                    lugar = row[3].value if len(row) > 3 else None
                    responsable = row[4].value if len(row) > 4 else None

                    # Validaciones
                    if not nombre or nombre.lower() == 'nombre':
                        continue  # Skip header row
                    
                    if not tipo or not fecha_str:
                        self.stdout.write(
                            self.style.WARNING(
                                f'Fila {row_idx}: Faltan tipo o fecha. Saltando.'
                            )
                        )
                        continue

                    # Normalizar tipo
                    tipo = str(tipo).lower().strip()
                    tipos_validos = ['reunion', 'taller', 'entrega_semillas', 'capacitacion', 'otro']
                    if tipo not in tipos_validos:
                        tipo = 'otro'

                    # Procesar fecha
                    if isinstance(fecha_str, str):
                        # Detectar formato de fecha
                        fecha = self._parse_fecha(fecha_str)
                    else:
                        fecha = fecha_str

                    if not fecha:
                        self.stdout.write(
                            self.style.WARNING(
                                f'Fila {row_idx}: No se pudo procesar fecha "{fecha_str}". Saltando.'
                            )
                        )
                        continue

                    nombre_limpio = str(nombre).strip()
                    lugar_limpio = str(lugar).strip() if lugar else '—'
                    responsable_limpio = str(responsable).strip() if responsable else '—'

                    # Agrupar por nombre de asunto
                    asuntos_map[nombre_limpio].append({
                        'tipo': tipo,
                        'fecha': fecha,
                        'lugar': lugar_limpio,
                        'responsable': responsable_limpio
                    })

                except Exception as e:
                    self.stdout.write(
                        self.style.WARNING(
                            f'Fila {row_idx}: Error procesando datos: {str(e)}'
                        )
                    )
                    continue

            # Crear asuntos y registros
            asuntos_creados = 0
            registros_creados = 0
            errores = []

            for nombre, registros in asuntos_map.items():
                try:
                    # Crear o actualizar asunto
                    asunto, creado = Asunto.objects.get_or_create(
                        nombre=nombre,
                        defaults={
                            'tipo': registros[0]['tipo'],
                            'total_instancias': len(registros),
                            'descripcion': f'Asunto importado desde Excel'
                        }
                    )

                    if creado:
                        asuntos_creados += 1
                        self.stdout.write(f'✓ Asunto creado: {nombre}')
                    else:
                        # Actualizar contador de instancias
                        asunto.total_instancias = len(registros)
                        asunto.save()
                        self.stdout.write(f'↻ Asunto existente: {nombre}')

                    # Crear registros (instancias) de este asunto
                    for registro_data in registros:
                        try:
                            registro, creado = RegistroAsunto.objects.get_or_create(
                                asunto=asunto,
                                fecha=registro_data['fecha'],
                                defaults={
                                    'lugar': registro_data['lugar'],
                                    'responsable': registro_data['responsable']
                                }
                            )
                            if creado:
                                registros_creados += 1
                        except Exception as e:
                            errores.append(f'Registro {nombre} ({registro_data["fecha"]}): {str(e)}')

                except Exception as e:
                    errores.append(f'Asunto {nombre}: {str(e)}')

            # Resumen
            self.stdout.write('\n' + '='*60)
            self.stdout.write(self.style.SUCCESS(f'✓ Importación completada'))
            self.stdout.write(f'  Asuntos creados: {asuntos_creados}')
            self.stdout.write(f'  Registros (instancias) creados: {registros_creados}')
            self.stdout.write(f'  Total asuntos en sistema: {Asunto.objects.count()}')
            self.stdout.write(f'  Total registros en sistema: {RegistroAsunto.objects.count()}')

            if errores:
                self.stdout.write(self.style.WARNING(f'\n⚠ {len(errores)} errores encontrados:'))
                for error in errores[:10]:  # Mostrar primeros 10 errores
                    self.stdout.write(f'  - {error}')
                if len(errores) > 10:
                    self.stdout.write(f'  ... y {len(errores) - 10} errores más')

        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'Error: Archivo no encontrado: {archivo}'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error: {str(e)}'))

    def _parse_fecha(self, fecha_str):
        """
        Intenta parsear la fecha en varios formatos comunes en español.
        """
        fecha_str = str(fecha_str).strip()
        
        # Formatos a intentar
        formatos = [
            '%d/%m/%Y',      # 19/11/2025
            '%d-%m-%Y',      # 19-11-2025
            '%Y-%m-%d',      # 2025-11-19
            '%d de %B %Y',   # 19 de November 2025 (con nombres en inglés)
            '%d de %b %Y',   # 19 de Nov 2025
        ]

        # Intentar con nombres de meses en español
        meses_es = {
            'enero': 'January',
            'febrero': 'February',
            'marzo': 'March',
            'abril': 'April',
            'mayo': 'May',
            'junio': 'June',
            'julio': 'July',
            'agosto': 'August',
            'septiembre': 'September',
            'octubre': 'October',
            'noviembre': 'November',
            'diciembre': 'December',
        }

        fecha_normalizada = fecha_str.lower()
        for mes_es, mes_en in meses_es.items():
            fecha_normalizada = fecha_normalizada.replace(mes_es, mes_en)

        # Intentar parsear
        for formato in formatos:
            try:
                return datetime.strptime(fecha_normalizada, formato).date()
            except ValueError:
                continue

        return None
