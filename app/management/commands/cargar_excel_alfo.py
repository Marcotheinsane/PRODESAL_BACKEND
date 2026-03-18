from django.core.management.base import BaseCommand
from app.models import (
    Clientes, Hogar, CargaFamiliar, InformacionSalud, 
    RedConocida, RegistroSocial, AnotacionesHogar
)
import openpyxl
from datetime import datetime
import os


class Command(BaseCommand):
    help = 'Actualiza fichas sociales desde "Sistematización Ficha Social (Alfo).xlsx" - Todas las personas ya existen'

    def add_arguments(self, parser):
        parser.add_argument(
            '--archivo',
            type=str,
            default='Sistematización Ficha Social (Alfo).xlsx',
            help='Ruta del archivo Excel'
        )
        parser.add_argument(
            '--hoja',
            type=str,
            default='Sx Fichas',
            help='Nombre de la hoja a cargar'
        )

    def _parse_date(self, fecha):
        """Parsear fecha de diferentes formatos"""
        if not fecha:
            return None
        if isinstance(fecha, datetime):
            return fecha.date()
        if isinstance(fecha, str):
            try:
                return datetime.strptime(fecha, '%d/%m/%Y').date()
            except:
                try:
                    return datetime.strptime(fecha, '%Y-%m-%d').date()
                except:
                    return None
        return None

    def handle(self, *args, **options):
        archivo = options['archivo']
        nombre_hoja = options['hoja']

        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR(f'❌ Archivo no encontrado: {archivo}'))
            return

        contador_actualizadas = 0
        contador_nuevos_hogares = 0
        contador_familiares = 0
        contador_no_encontradas = 0

        try:
            wb = openpyxl.load_workbook(archivo)
            
            if nombre_hoja not in wb.sheetnames:
                self.stdout.write(self.style.ERROR(f'❌ Hoja "{nombre_hoja}" no existe'))
                return

            ws = wb[nombre_hoja]
            self.stdout.write(self.style.SUCCESS(f'📄 Leyendo hoja: {nombre_hoja}'))

            # Posiciones de columnas
            col_n = 1
            col_fecha = 2
            col_nombres = 3
            col_apellidos = 4
            col_rut = 5
            col_edad = 6
            col_num_doc = 7
            col_fecha_nac = 8
            col_sector = 9
            col_estado_civil = 10
            col_escolaridad = 11
            col_telefono = 12
            col_tipo_tenencia = 13
            col_saneada = 14
            col_ingresos = 15
            
            # Información Familiar
            col_parentesco = 16
            col_nombre_familiar = 17
            col_edad_familiar = 18
            col_rut_familiar = 19
            col_ocupacion_familiar = 20
            col_escolaridad_familiar = 21
            
            # Registro Social
            col_posee_registro = 22
            col_puntaje = 23
            col_actualizado = 24
            
            # Salud
            col_enfermedad = 25
            col_medicamentos = 26
            col_controles = 27
            
            # Redes
            col_redes_conoce = 28
            col_redes_participa = 29
            col_redes_apoyo = 30
            
            # Otros
            col_necesidades = 31
            col_observaciones = 32
            col_recomendaciones = 33
            col_luz = 34
            col_agua = 35
            col_rol = 36

            hogar_actual = None
            
            # Iterar sobre filas (desde fila 3: fila 1-2 son headers)
            for row_num in range(3, ws.max_row + 1):
                try:
                    # Leer N° para detectar si es persona principal
                    n_orden = ws.cell(row=row_num, column=col_n).value
                    
                    # Si hay N°, es persona principal (nuevo hogar)
                    if n_orden is not None:
                        # Buscar persona por RUT
                        rut = ws.cell(row=row_num, column=col_rut).value
                        if not rut:
                            continue
                        
                        rut_str = str(rut).strip()
                        
                        # Buscar cliente existente
                        try:
                            cliente = Clientes.objects.get(rut=rut_str)
                        except Clientes.DoesNotExist:
                            contador_no_encontradas += 1
                            self.stdout.write(self.style.WARNING(f'⚠️  Fila {row_num}: RUT no encontrado en BD: {rut_str}'))
                            continue
                        
                        # Actualizar información personal
                        edad = ws.cell(row=row_num, column=col_edad).value
                        cliente.edad = int(edad) if edad else cliente.edad
                        
                        fecha_nac = self._parse_date(ws.cell(row=row_num, column=col_fecha_nac).value)
                        if fecha_nac:
                            cliente.fecha_nacimiento = fecha_nac
                        
                        cliente.numero_documento = str(ws.cell(row=row_num, column=col_num_doc).value or '').strip() or None
                        cliente.sector = str(ws.cell(row=row_num, column=col_sector).value or '').strip() or cliente.sector
                        cliente.estado_civil = str(ws.cell(row=row_num, column=col_estado_civil).value or '').strip().lower() or cliente.estado_civil
                        cliente.escolaridad = self._normalizar_escolaridad(ws.cell(row=row_num, column=col_escolaridad).value) or cliente.escolaridad
                        cliente.telefono = str(ws.cell(row=row_num, column=col_telefono).value or '').strip() or cliente.telefono
                        cliente.tipo_tenencia = self._normalizar_tenencia(ws.cell(row=row_num, column=col_tipo_tenencia).value) or cliente.tipo_tenencia
                        
                        saneada = ws.cell(row=row_num, column=col_saneada).value
                        if self._parse_bool(saneada) is not None:
                            cliente.vivienda_saneada = self._parse_bool(saneada)
                        
                        cliente.ingresos = str(ws.cell(row=row_num, column=col_ingresos).value or '').strip() or cliente.ingresos
                        cliente.save()
                        
                        # Crear o actualizar hogar
                        hogar_actual, creado = Hogar.objects.get_or_create(
                            persona_principal=cliente,
                            defaults={
                                'luz': self._parse_bool(ws.cell(row=row_num, column=col_luz).value),
                                'agua': self._parse_bool(ws.cell(row=row_num, column=col_agua).value),
                                'rol': str(ws.cell(row=row_num, column=col_rol).value or '').strip() or None,
                                'fecha_registro': self._parse_date(ws.cell(row=row_num, column=col_fecha).value) or datetime.now().date(),
                            }
                        )
                        
                        if creado:
                            contador_nuevos_hogares += 1
                        
                        # Actualizar hogar existente
                        if not creado:
                            luz = self._parse_bool(ws.cell(row=row_num, column=col_luz).value)
                            if luz is not None:
                                hogar_actual.luz = luz
                            agua = self._parse_bool(ws.cell(row=row_num, column=col_agua).value)
                            if agua is not None:
                                hogar_actual.agua = agua
                            hogar_actual.rol = str(ws.cell(row=row_num, column=col_rol).value or '').strip() or hogar_actual.rol
                            hogar_actual.save()
                        
                        # Crear o actualizar información de salud
                        salud, _ = InformacionSalud.objects.get_or_create(
                            persona=cliente,
                            defaults={
                                'enfermedad': str(ws.cell(row=row_num, column=col_enfermedad).value or '').strip() or None,
                                'medicamentos': str(ws.cell(row=row_num, column=col_medicamentos).value or '').strip() or None,
                                'controles_medicos': str(ws.cell(row=row_num, column=col_controles).value or '').strip() or None,
                                'toma_medicamentos': self._parse_bool(ws.cell(row=row_num, column=col_medicamentos).value),
                            }
                        )
                        
                        salud.enfermedad = str(ws.cell(row=row_num, column=col_enfermedad).value or '').strip() or salud.enfermedad
                        toma_med = self._parse_bool(ws.cell(row=row_num, column=col_medicamentos).value)
                        if toma_med is not None:
                            salud.toma_medicamentos = toma_med
                        salud.save()
                        
                        # Crear o actualizar redes
                        redes, _ = RedConocida.objects.get_or_create(
                            persona=cliente,
                            defaults={
                                'redes_conoce': str(ws.cell(row=row_num, column=col_redes_conoce).value or '').strip() or None,
                                'redes_participa': str(ws.cell(row=row_num, column=col_redes_participa).value or '').strip() or None,
                                'redes_apoyo': str(ws.cell(row=row_num, column=col_redes_apoyo).value or '').strip() or None,
                            }
                        )
                        
                        redes.redes_conoce = str(ws.cell(row=row_num, column=col_redes_conoce).value or '').strip() or redes.redes_conoce
                        redes.redes_participa = str(ws.cell(row=row_num, column=col_redes_participa).value or '').strip() or redes.redes_participa
                        redes.redes_apoyo = str(ws.cell(row=row_num, column=col_redes_apoyo).value or '').strip() or redes.redes_apoyo
                        redes.save()
                        
                        # Crear o actualizar registro social
                        reg_social, _ = RegistroSocial.objects.get_or_create(
                            persona=cliente,
                            defaults={
                                'posee_registro': self._parse_bool(ws.cell(row=row_num, column=col_posee_registro).value),
                                'puntaje': str(ws.cell(row=row_num, column=col_puntaje).value or '').strip() or None,
                                'actualizado': self._parse_bool(ws.cell(row=row_num, column=col_actualizado).value),
                            }
                        )
                        
                        posee = self._parse_bool(ws.cell(row=row_num, column=col_posee_registro).value)
                        if posee is not None:
                            reg_social.posee_registro = posee
                        reg_social.puntaje = str(ws.cell(row=row_num, column=col_puntaje).value or '').strip() or reg_social.puntaje
                        act = self._parse_bool(ws.cell(row=row_num, column=col_actualizado).value)
                        if act is not None:
                            reg_social.actualizado = act
                        reg_social.save()
                        
                        # Crear o actualizar anotaciones
                        anotaciones, _ = AnotacionesHogar.objects.get_or_create(
                            hogar=hogar_actual,
                            defaults={
                                'necesidades': str(ws.cell(row=row_num, column=col_necesidades).value or '').strip() or None,
                                'observaciones': str(ws.cell(row=row_num, column=col_observaciones).value or '').strip() or None,
                                'recomendaciones': str(ws.cell(row=row_num, column=col_recomendaciones).value or '').strip() or None,
                            }
                        )
                        
                        anotaciones.necesidades = str(ws.cell(row=row_num, column=col_necesidades).value or '').strip() or anotaciones.necesidades
                        anotaciones.observaciones = str(ws.cell(row=row_num, column=col_observaciones).value or '').strip() or anotaciones.observaciones
                        anotaciones.recomendaciones = str(ws.cell(row=row_num, column=col_recomendaciones).value or '').strip() or anotaciones.recomendaciones
                        anotaciones.save()
                        
                        contador_actualizadas += 1
                        self.stdout.write(self.style.SUCCESS(f'✅ Fila {row_num}: Actualizado {cliente.nombres} ({rut_str})'))
                    
                    # Si NO hay N°, es carga familiar del hogar anterior
                    elif hogar_actual is not None:
                        nombre_familiar = ws.cell(row=row_num, column=col_nombre_familiar).value
                        if not nombre_familiar:
                            continue
                        
                        parentesco_val = ws.cell(row=row_num, column=col_parentesco).value
                        parentesco = self._normalizar_parentesco(parentesco_val)
                        
                        # Crear o actualizar carga familiar
                        rut_familiar = str(ws.cell(row=row_num, column=col_rut_familiar).value or '').strip() or None
                        
                        carga, creado = CargaFamiliar.objects.get_or_create(
                            hogar=hogar_actual,
                            nombre=str(nombre_familiar).strip(),
                            defaults={
                                'parentesco': parentesco or 'otros',
                                'edad': int(ws.cell(row=row_num, column=col_edad_familiar).value) if ws.cell(row=row_num, column=col_edad_familiar).value else None,
                                'rut': rut_familiar,
                                'ocupacion': str(ws.cell(row=row_num, column=col_ocupacion_familiar).value or '').strip() or None,
                                'escolaridad': self._normalizar_escolaridad(ws.cell(row=row_num, column=col_escolaridad_familiar).value),
                            }
                        )
                        
                        if not creado:
                            carga.parentesco = parentesco or 'otros'
                            edad_fam = ws.cell(row=row_num, column=col_edad_familiar).value
                            if edad_fam:
                                carga.edad = int(edad_fam)
                            carga.rut = rut_familiar
                            carga.ocupacion = str(ws.cell(row=row_num, column=col_ocupacion_familiar).value or '').strip() or carga.ocupacion
                            carga.save()
                        
                        contador_familiares += 1
                        self.stdout.write(self.style.SUCCESS(f'  └─ {nombre_familiar} ({parentesco})'))
                
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'❌ Fila {row_num}: {str(e)}'))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Error al procesar archivo: {e}'))
            return

        # Resumen
        self.stdout.write(
            self.style.SUCCESS(
                f'\n✅ ACTUALIZACIÓN COMPLETADA:\n'
                f'   Personas actualizadas: {contador_actualizadas}\n'
                f'   Nuevos hogares: {contador_nuevos_hogares}\n'
                f'   Carga Familiar: {contador_familiares}\n'
                f'   No encontradas: {contador_no_encontradas}'
            )
        )

    @staticmethod
    def _parse_bool(valor):
        """Parsear valor a booleano"""
        if valor is None:
            return None
        valor_str = str(valor).strip().upper()
        if valor_str in ['SI', 'YES', 'TRUE', '1']:
            return True
        elif valor_str in ['NO', 'FALSE', '0']:
            return False
        return None

    @staticmethod
    def _normalizar_escolaridad(valor):
        """Normalizar escolaridad a opciones del modelo"""
        if not valor:
            return None
        valor_str = str(valor).strip().upper()
        
        mapping = {
            'SIN ESTUDIOS': 'sin_estudios',
            'BASICA INCOMPLETA': 'basica_incompleta',
            'BASICA COMPLETA': 'basica_completa',
            'BASICA COMP': 'basica_completa',
            'BASICA INCOM': 'basica_incompleta',
            'MEDIA INCOMPLETA': 'media_incompleta',
            'MEDIA COMPLETA': 'media_completa',
            'MEDIA COMP': 'media_completa',
            'MEDIA INCOM': 'media_incompleta',
            'TECNICA': 'tecnica',
            'SUPERIOR': 'superior',
        }
        
        return mapping.get(valor_str, None)

    @staticmethod
    def _normalizar_tenencia(valor):
        """Normalizar tipo de tenencia"""
        if not valor:
            return None
        valor_str = str(valor).strip().upper()
        
        mapping = {
            'PROPIETARIO': 'propietario',
            'PROPIETARIA': 'propietario',
            'ARRENDATARIO': 'arrendatario',
            'ARRENDATARIA': 'arrendatario',
            'PRESTADO': 'prestado',
            'PRESTADA': 'prestado',
            'OCUPANTE': 'ocupante',
        }
        
        return mapping.get(valor_str, None)

    @staticmethod
    def _normalizar_parentesco(valor):
        """Normalizar parentesco"""
        if not valor:
            return 'otros'
        valor_str = str(valor).strip().upper()
        
        mapping = {
            'HIJO': 'hijo',
            'HIJA': 'hijo',
            'PADRE': 'padre',
            'MADRE': 'madre',
            'ABUELO': 'abuelo',
            'ABUELA': 'abuelo',
            'HERMANO': 'hermano',
            'HERMANA': 'hermano',
        }
        
        return mapping.get(valor_str, 'otros')
